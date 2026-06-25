# -*- coding: utf-8 -*-
import argparse
import os
import re
from datetime import date, datetime, time
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = os.getenv("IGCOM_BASE_DIR", "/igcom")
OUTPUT_ROOT_DIR = os.getenv("IGCOM_OUTPUT_ROOT_DIR", os.path.join(BASE_DIR, "outputs"))
MONTHLY_REPORT_DIR = os.getenv(
    "MONTHLY_REPORT_DIR",
    os.path.join(OUTPUT_ROOT_DIR, "monthly_reports"),
)
DEFAULT_TEMPLATE_FILE = os.getenv(
    "DETAILED_REPORT_TEMPLATE_FILE",
    str(Path(__file__).with_name("IG变化率表格(英区).xlsx")),
)

DETAILED_FILE_RE = re.compile(r"^All_Products_Full_1h_30min_(\d{8})\.xlsx$")

DATA_SHEET_NAME = "All_Full_Data"
DATETIME_COL = "DateTime (London)"
CLOSE_COL = "Close"
PRODUCT_COL = "Product Name"

PRODUCT_LABELS = [
    ("US500", lambda name: "US 500" in name),
    ("HK50", lambda name: "Hong Kong" in name),
    ("NIKKEI", lambda name: "Japan 225" in name),
    ("USD/JPY", lambda name: name == "USD/JPY"),
    ("USD/SGD", lambda name: name == "USD/SGD"),
    ("UK100", lambda name: "UK 100" in name),
    ("GBP/USD", lambda name: name == "GBP/USD"),
    ("France40", lambda name: "France 40" in name),
    ("EUR/USD", lambda name: name == "EUR/USD"),
    ("USD/INR", lambda name: name.startswith("USD/INR")),
    ("Germany40", lambda name: "Germany 40" in name),
    ("USD/CNH", lambda name: name == "USD/CNH"),
    ("USD/TWD", lambda name: name.startswith("USD/TWD")),
    ("Australia200", lambda name: "Australia 200" in name),
    ("AUD/USD", lambda name: name == "AUD/USD"),
    ("USDKRW", lambda name: name.startswith("USD/KRW")),
    ("USDMXN", lambda name: name == "USD/MXN"),
]

CLOSE_COL_PRODUCT = {
    2: "US500",
    4: "HK50",
    7: "NIKKEI",
    10: "USD/JPY",
    14: "USD/SGD",
    16: "UK100",
    19: "GBP/USD",
    22: "France40",
    25: "EUR/USD",
    28: "USD/INR",
    30: "Germany40",
    34: "USD/CNH",
    38: "USD/TWD",
    40: "Australia200",
    43: "AUD/USD",
    46: "USDKRW",
    48: "USDMXN",
}

DIRECT_CHANGE_BY_CLOSE = {close_col: close_col + 1 for close_col in CLOSE_COL_PRODUCT}

DERIVED_FORMULAS = {
    6: lambda row: f'=IFERROR(E{row}-C{row},"")',
    9: lambda row: f'=IFERROR(H{row}-C{row},"")',
    12: lambda row: f'=IFERROR(H{row}-K{row},"")',
    13: lambda row: f'=IFERROR(L{row}-C{row},"")',
    18: lambda row: f'=IFERROR(Q{row}-C{row},"")',
    21: lambda row: f'=IFERROR(R{row}+T{row},"")',
    24: lambda row: f'=IFERROR(W{row}-C{row},"")',
    27: lambda row: f'=IFERROR(X{row}+Z{row},"")',
    32: lambda row: f'=IFERROR(AE{row}-C{row},"")',
    33: lambda row: f'=IFERROR(AF{row}+Z{row},"")',
    36: lambda row: f'=IFERROR(AI{row}-K{row},"")',
    37: lambda row: f'=IFERROR(AI{row}-O{row},"")',
    42: lambda row: f'=IFERROR(AO{row}-C{row},"")',
    45: lambda row: f'=IFERROR(AP{row}+AR{row},"")',
}

HEADER_DARK = "17365D"
HEADER_LIGHT = "EAF2F8"
GRID = "B7C4D0"
GRID_STRONG = "5B6B7D"
ROW_ALT = "F8FBFD"
BLANK_FILL = "EEF2F6"
TEXT_DARK = "1F2937"
TEXT_NAVY = "17365D"


def collect_detailed_workbooks(output_root_dir: str) -> List[Path]:
    """Find generated full 1h/30Min data workbooks under the outputs directory."""
    root = Path(output_root_dir)
    if not root.exists():
        return []

    candidates = []
    for path in root.rglob("All_Products_Full_1h_30min_*.xlsx"):
        if DETAILED_FILE_RE.match(path.name):
            candidates.append(path)

    return sorted(set(candidates), key=lambda item: (item.stat().st_mtime, str(item)))


def _parse_report_month(report_month: Optional[str]) -> str:
    if report_month:
        month = str(report_month)
    else:
        month = datetime.now().strftime("%Y%m")
    if not re.match(r"^\d{6}$", month):
        raise ValueError("report_month must use YYYYMM format")
    return month


def _date_from_detailed_file(path: Path) -> date:
    match = DETAILED_FILE_RE.match(path.name)
    if not match:
        raise RuntimeError(f"Detailed workbook name must include YYYYMMDD: {path}")
    return datetime.strptime(match.group(1), "%Y%m%d").date()


def _label_product(name) -> Optional[str]:
    text = str(name or "").strip()
    for label, predicate in PRODUCT_LABELS:
        if predicate(text):
            return label
    return None


def _coerce_datetime(value) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
    return None


def _safe_float(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _extract_records(input_files: Sequence[Path]):
    records: Dict[Tuple[date, str, str], Optional[float]] = {}
    all_dates = set()

    for path in input_files:
        all_dates.add(_date_from_detailed_file(path))
        wb = load_workbook(path, data_only=True, read_only=True)
        try:
            if DATA_SHEET_NAME not in wb.sheetnames:
                raise RuntimeError(f"Workbook missing {DATA_SHEET_NAME} sheet: {path}")
            ws = wb[DATA_SHEET_NAME]
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            header_idx = {name: idx for idx, name in enumerate(header)}
            required = [DATETIME_COL, CLOSE_COL, PRODUCT_COL]
            missing = [name for name in required if name not in header_idx]
            if missing:
                raise RuntimeError(f"Workbook missing required columns {missing}: {path}")

            for row in ws.iter_rows(min_row=2, values_only=True):
                dt = _coerce_datetime(row[header_idx[DATETIME_COL]])
                product = _label_product(row[header_idx[PRODUCT_COL]])
                if dt is None or product is None:
                    continue
                current_date = dt.date()
                time_text = dt.strftime("%H:%M")
                all_dates.add(current_date)
                records[(current_date, time_text, product)] = _safe_float(row[header_idx[CLOSE_COL]])
        finally:
            wb.close()

    return records, sorted(all_dates)


def _visible_dates_for_month(all_dates, report_month: str):
    year = int(report_month[:4])
    month = int(report_month[4:])
    month_dates = [item for item in all_dates if item.year == year and item.month == month]
    if not month_dates:
        raise RuntimeError(f"No detailed data found for report month {report_month}")
    return month_dates


def _load_headers(template_file: str):
    path = Path(template_file)
    if not path.exists():
        raise RuntimeError(f"Template workbook not found: {path}")

    wb = load_workbook(path, data_only=False, read_only=True)
    try:
        ws = wb.worksheets[0]
        headers = []
        for row in range(1, 3):
            headers.append([ws.cell(row, col).value for col in range(1, 50)])
        return headers
    finally:
        wb.close()


def _time_points():
    return [f"{hour:02d}:{minute:02d}" for hour in range(24) for minute in (0, 30)]


def _sheet_name(time_text: str) -> str:
    return time_text.replace(":", "_")


def _row_label(date_obj: date, time_text: str) -> str:
    return f"{date_obj.strftime('%Y/%m/%d')}-{time_text}时Close"


def _write_formulas(ws, max_row: int) -> None:
    first_data_row = 3
    for row in range(first_data_row, max_row + 1):
        for close_col, change_col in DIRECT_CHANGE_BY_CLOSE.items():
            cell = ws.cell(row, change_col)
            if row == first_data_row:
                cell.value = None
            else:
                close_letter = get_column_letter(close_col)
                cell.value = f'=IFERROR({close_letter}{row}/{close_letter}{row - 1}-1,"")'

        for derived_col, formula_fn in DERIVED_FORMULAS.items():
            cell = ws.cell(row, derived_col)
            cell.value = None if row == first_data_row else formula_fn(row)


def _apply_style(ws, max_row: int, max_col: int = 49) -> None:
    thin = Side(style="thin", color=GRID)
    medium = Side(style="medium", color=GRID_STRONG)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_border = Border(left=thin, right=thin, top=medium, bottom=medium)

    ws.freeze_panes = "B3"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 38
    ws.row_dimensions[2].height = 26
    ws.column_dimensions["A"].width = 26

    change_cols = set(DIRECT_CHANGE_BY_CLOSE.values()) | set(DERIVED_FORMULAS)
    close_cols = set(CLOSE_COL_PRODUCT)

    for col in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 11 if col in close_cols else 10

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            cell.border = header_border if row <= 2 else border
            cell.alignment = Alignment(
                horizontal="center" if row <= 2 else ("left" if col == 1 else "right"),
                vertical="center",
                wrap_text=row <= 2,
            )
            cell.font = Font(
                name="Aptos",
                size=10,
                bold=row <= 2 or col == 1,
                color=("FFFFFF" if row == 1 or (col == 1 and row <= 2) else (TEXT_NAVY if row <= 2 or col == 1 else TEXT_DARK)),
            )
            if row == 1 or (col == 1 and row <= 2):
                cell.fill = PatternFill("solid", fgColor=HEADER_DARK)
            elif row == 2:
                cell.fill = PatternFill("solid", fgColor=HEADER_LIGHT)
            elif row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ROW_ALT)

            if row >= 3 and col in close_cols:
                cell.number_format = "0.0000"
            elif row >= 3 and col in change_cols:
                cell.number_format = "0.00%"


def _shade_blank_cells(ws, max_row: int) -> None:
    change_cols = set(DIRECT_CHANGE_BY_CLOSE.values()) | set(DERIVED_FORMULAS)
    close_cols = set(CLOSE_COL_PRODUCT)
    blank_fill = PatternFill("solid", fgColor=BLANK_FILL)

    for row in range(3, max_row + 1):
        if row == 3:
            for col in change_cols:
                ws.cell(row, col).fill = blank_fill
        for col in close_cols:
            if ws.cell(row, col).value is None:
                ws.cell(row, col).fill = blank_fill


def build_detailed_monthly_report(
    output_root_dir: str = OUTPUT_ROOT_DIR,
    report_dir: str = MONTHLY_REPORT_DIR,
    output_file: Optional[str] = None,
    report_month: Optional[str] = None,
    input_files: Optional[Iterable[str]] = None,
    template_file: str = DEFAULT_TEMPLATE_FILE,
) -> str:
    """Build the monthly detailed 48-half-hour formula workbook from full 1h/30Min data."""
    month = _parse_report_month(report_month)
    if input_files is None:
        files = collect_detailed_workbooks(output_root_dir)
    else:
        files = [Path(path) for path in input_files]

    if not files:
        raise RuntimeError("No detailed 1h/30Min workbooks found")

    files = sorted(files, key=lambda item: (item.stat().st_mtime, str(item)))
    headers = _load_headers(template_file)
    records, all_dates = _extract_records(files)
    visible_dates = _visible_dates_for_month(all_dates, month)

    wb = Workbook()
    wb.remove(wb.active)
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    for time_text in _time_points():
        ws = wb.create_sheet(_sheet_name(time_text))
        ws.append(headers[0])
        ws.append(headers[1])

        for date_obj in visible_dates:
            row_values = [None] * 49
            row_values[0] = _row_label(date_obj, time_text)
            for col, product in CLOSE_COL_PRODUCT.items():
                row_values[col - 1] = records.get((date_obj, time_text, product))
            ws.append(row_values)

        max_row = len(visible_dates) + 2
        _write_formulas(ws, max_row)
        _apply_style(ws, max_row)
        _shade_blank_cells(ws, max_row)

    if output_file is None:
        output_file = os.path.join(report_dir, f"IG变化率_{month}_详细版.xlsx")

    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.active = 0
    wb.save(output_path)
    wb.close()
    return str(output_path)


def build_and_send_detailed_monthly_report(
    output_root_dir: str = OUTPUT_ROOT_DIR,
    report_dir: str = MONTHLY_REPORT_DIR,
    report_month: Optional[str] = None,
    send_email: bool = True,
    template_file: str = DEFAULT_TEMPLATE_FILE,
) -> str:
    month = _parse_report_month(report_month)
    output_file = os.path.join(report_dir, f"IG变化率_{month}_详细版.xlsx")
    report_file = build_detailed_monthly_report(
        output_root_dir=output_root_dir,
        report_dir=report_dir,
        output_file=output_file,
        report_month=month,
        template_file=template_file,
    )

    if send_email:
        from config import get_gmail_config
        from monthly_report import send_gmail_with_attachments

        gmail_config = get_gmail_config()
        sent = send_gmail_with_attachments(
            send_usr=gmail_config.send_usr,
            send_pwd=gmail_config.send_pwd,
            receive_usr_list=gmail_config.receive_usr_list,
            attachment_paths=[report_file],
            email_title=f"IG变化率详细版 - {month}",
            content="这是当天生成的IG变化率月度详细版，请查收。",
            email_server=gmail_config.email_server,
            email_port=gmail_config.email_port,
        )
        if not sent:
            raise RuntimeError("Detailed monthly report was generated, but email sending failed")

    return report_file


def main() -> None:
    parser = argparse.ArgumentParser(description="Build IG detailed monthly 48-half-hour report.")
    parser.add_argument("--report-month", dest="report_month", help="Report month in YYYYMM format.")
    parser.add_argument("--output-root-dir", default=OUTPUT_ROOT_DIR, help="Root directory containing historical_data_* folders.")
    parser.add_argument("--report-dir", default=MONTHLY_REPORT_DIR, help="Directory where monthly reports are saved.")
    parser.add_argument("--template-file", default=DEFAULT_TEMPLATE_FILE, help="Workbook used for the two header rows.")
    args = parser.parse_args()

    send_email_value = os.getenv(
        "DETAILED_MONTHLY_REPORT_SEND_EMAIL",
        os.getenv("MONTHLY_REPORT_SEND_EMAIL", "true"),
    ).strip().lower()
    send_email = send_email_value not in {"0", "false", "no", "off"}
    report_file = build_and_send_detailed_monthly_report(
        output_root_dir=args.output_root_dir,
        report_dir=args.report_dir,
        report_month=args.report_month,
        send_email=send_email,
        template_file=args.template_file,
    )
    print(f"Detailed monthly report ready: {report_file}")


if __name__ == "__main__":
    main()
