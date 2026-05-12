# -*- coding: utf-8 -*-
import os
import re
from copy import copy
from datetime import datetime
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from smtplib import SMTP
from typing import Dict, Iterable, List, Optional, Sequence

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter


BASE_DIR = os.getenv("IGCOM_BASE_DIR", "/igcom")
OUTPUT_ROOT_DIR = os.getenv("IGCOM_OUTPUT_ROOT_DIR", os.path.join(BASE_DIR, "outputs"))
MONTHLY_REPORT_DIR = os.getenv(
    "MONTHLY_REPORT_DIR",
    os.path.join(OUTPUT_ROOT_DIR, "monthly_reports"),
)

TIME_SHEET_IDXS = [0, 1, 2, 3]
CHANGE_SHEET_OFFSET = 4
BASE_TIMES = {0: "05", 1: "07", 2: "15"}
TIME_ROWS = {
    0: ["05", "18", "18:30", "19", "19", "20"],
    1: ["07", "18", "18:30", "19", "19", "20"],
    2: ["15", "18", "18:30", "19", "19", "20"],
    3: ["18", "18", "18:30", "19", "19", "20"],
}
PREV_REF_TIME_20 = {0: "19", 1: "20", 2: "20", 3: "19", 4: "20", 5: "20"}

DAILY_FILE_RE = re.compile(r"^IG变化率_(\d{8})\.xlsx$")
LABEL_RE = re.compile(r"^(\d{4}/\d{2}/\d{2})-(.+?)Close$")
FORMULA_TWO_REF_RE = re.compile(r"^=\$?([A-Z]+)\$?\d+([+-])\$?([A-Z]+)\$?\d+$")
ROW_REF_RE = re.compile(r"(\$?[A-Z]+\$?)\d+")


def collect_daily_workbooks(output_root_dir: str) -> List[Path]:
    """Find generated daily workbooks under the outputs directory."""
    root = Path(output_root_dir)
    if not root.exists():
        return []

    candidates = []
    for path in root.rglob("IG变化率_*.xlsx"):
        if DAILY_FILE_RE.match(path.name):
            candidates.append(path)

    return sorted(set(candidates), key=lambda item: (item.stat().st_mtime, str(item)))


def _copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst._style = copy(src._style)
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)


def _apply_row_style(ws, row_style_refs, row_idx: int, row_type_idx: int) -> None:
    style_ref = row_style_refs[row_type_idx]
    ws.row_dimensions[row_idx].height = style_ref["height"]
    for col, src_cell in enumerate(style_ref["cells"], start=1):
        _copy_cell_style(src_cell, ws.cell(row_idx, col))


def _parse_report_month(report_month: Optional[str]) -> str:
    if report_month:
        month = str(report_month)
    else:
        month = datetime.now().strftime("%Y%m")
    if not re.match(r"^\d{6}$", month):
        raise ValueError("report_month must use YYYYMM format")
    return month


def _same_row_formula(formula, row_num: int):
    if not isinstance(formula, str) or not formula.startswith("="):
        return None
    return ROW_REF_RE.sub(lambda match: f"{match.group(1)}{row_num}", formula)


def _extract_sheet_info(style_wb):
    sheet_info = {}
    for idx in TIME_SHEET_IDXS:
        ws = style_wb.worksheets[idx]
        close_cols = []
        direct_change_by_close = {}
        time_change_col_by_header = {}
        derived_cols = []

        for col in range(1, ws.max_column + 1):
            header = ws.cell(1, col).value
            subheader = ws.cell(2, col).value
            if subheader == "Close":
                close_cols.append(col)
                next_col = col + 1
                if next_col <= ws.max_column and ws.cell(2, next_col).value == "Change":
                    direct_change_by_close[col] = next_col
                    if header:
                        time_change_col_by_header[header] = next_col
            elif subheader == "Change" and header:
                derived_cols.append(col)
                time_change_col_by_header[header] = col

        sheet_info[idx] = {
            "close_cols": close_cols,
            "direct_change_by_close": direct_change_by_close,
            "derived_cols": derived_cols,
            "time_change_col_by_header": time_change_col_by_header,
        }

    return sheet_info


def _extract_records(input_files: Sequence[Path], sheet_info):
    records = {idx: {} for idx in TIME_SHEET_IDXS}
    all_dates = set()

    for path in input_files:
        wb = load_workbook(path, data_only=False, read_only=True)
        try:
            for idx in TIME_SHEET_IDXS:
                ws = wb.worksheets[idx]
                for row in range(3, ws.max_row + 1):
                    raw_label = ws.cell(row, 1).value
                    if not raw_label:
                        continue
                    match = LABEL_RE.match(str(raw_label).strip())
                    if not match:
                        continue

                    date_obj = datetime.strptime(match.group(1), "%Y/%m/%d").date()
                    time_part = match.group(2).replace("时", "").strip()
                    all_dates.add(date_obj)
                    records[idx].setdefault(date_obj, {}).setdefault(time_part, {})

                    for col in sheet_info[idx]["close_cols"]:
                        records[idx][date_obj][time_part][col] = ws.cell(row, col).value
        finally:
            wb.close()

    return records, sorted(all_dates)


def _visible_dates_for_month(all_dates, report_month: str):
    year = int(report_month[:4])
    month = int(report_month[4:])
    month_dates = [item for item in all_dates if item.year == year and item.month == month]
    if not month_dates:
        raise RuntimeError(f"No data found for report month {report_month}")

    first_month_date = month_dates[0]
    previous_dates = [item for item in all_dates if item < first_month_date]
    if previous_dates:
        return [previous_dates[-1]] + month_dates
    return month_dates


def _build_formula_refs_sheet(wb, sheet_info, records, visible_dates, prev_date_by_date):
    if "FormulaRefs" in wb.sheetnames:
        del wb["FormulaRefs"]

    ref_ws = wb.create_sheet("FormulaRefs")
    ref_ws.sheet_state = "hidden"
    ref_ws.cell(1, 1).value = "Key"
    for col in range(2, 50):
        ref_ws.cell(1, col).value = (
            wb.worksheets[0].cell(1, col).value or wb.worksheets[0].cell(2, col).value
        )

    ref_row_by_key: Dict[str, int] = {}

    def close_value(sheet_idx: int, date_obj, time_str: str, close_col: int):
        return records.get(sheet_idx, {}).get(date_obj, {}).get(time_str, {}).get(close_col)

    def add_ref_row(sheet_idx: int, date_obj, time_str: str) -> int:
        key = f"{wb.worksheets[sheet_idx].title}|{date_obj.strftime('%Y/%m/%d')}|{time_str}"
        if key in ref_row_by_key:
            return ref_row_by_key[key]

        row = len(ref_row_by_key) + 2
        ref_row_by_key[key] = row
        ref_ws.cell(row, 1).value = key
        for col in sheet_info[sheet_idx]["close_cols"]:
            ref_ws.cell(row, col).value = close_value(sheet_idx, date_obj, time_str, col)
        return row

    first_prev = prev_date_by_date.get(visible_dates[0])
    if first_prev:
        for idx in [0, 1, 2]:
            add_ref_row(idx, first_prev, BASE_TIMES[idx])
        add_ref_row(3, first_prev, "19")
        add_ref_row(3, first_prev, "20")

    return add_ref_row


def build_monthly_formula_report(
    output_root_dir: str = OUTPUT_ROOT_DIR,
    output_file: Optional[str] = None,
    report_month: Optional[str] = None,
    input_files: Optional[Iterable[str]] = None,
) -> str:
    """Build the monthly cumulative formula workbook from generated daily workbooks."""
    month = _parse_report_month(report_month)
    if input_files is None:
        files = collect_daily_workbooks(output_root_dir)
    else:
        files = [Path(path) for path in input_files]

    if not files:
        raise RuntimeError("No daily workbooks found")

    files = sorted(files, key=lambda item: (item.stat().st_mtime, str(item)))
    style_wb = load_workbook(files[-1], data_only=False)
    out_wb = load_workbook(files[-1], data_only=False)
    try:
        out_wb.calculation.fullCalcOnLoad = True
        out_wb.calculation.forceFullCalc = True
        out_wb.calculation.calcMode = "auto"
    except Exception:
        pass

    sheet_info = _extract_sheet_info(style_wb)
    records, all_dates = _extract_records(files, sheet_info)
    visible_dates = _visible_dates_for_month(all_dates, month)
    prev_date_by_date = {
        all_dates[idx]: all_dates[idx - 1]
        for idx in range(1, len(all_dates))
    }

    row_styles = {}
    for sheet_idx, ws in enumerate(style_wb.worksheets[:8]):
        row_styles[sheet_idx] = []
        for offset in range(6):
            ref_row = 9 + offset
            row_styles[sheet_idx].append({
                "height": ws.row_dimensions[ref_row].height,
                "cells": [ws.cell(ref_row, col) for col in range(1, ws.max_column + 1)],
            })

    for ws in out_wb.worksheets:
        if ws.max_row > 2:
            ws.delete_rows(3, ws.max_row - 2)

    add_ref_row = _build_formula_refs_sheet(
        out_wb,
        sheet_info,
        records,
        visible_dates,
        prev_date_by_date,
    )

    def close_value(sheet_idx: int, date_obj, time_str: str, close_col: int):
        return records.get(sheet_idx, {}).get(date_obj, {}).get(time_str, {}).get(close_col)

    def denominator_ref(sheet_idx: int, date_index: int, row_type_idx: int, close_col: int, block_start: int):
        if sheet_idx in BASE_TIMES:
            if row_type_idx == 0:
                if date_index == 0:
                    previous = prev_date_by_date.get(visible_dates[date_index])
                    if not previous:
                        return None
                    ref_row = add_ref_row(sheet_idx, previous, BASE_TIMES[sheet_idx])
                    return f"FormulaRefs!{get_column_letter(close_col)}{ref_row}"
                return f"{get_column_letter(close_col)}{block_start - 6}"
            return f"${get_column_letter(close_col)}${block_start}"

        if date_index == 0:
            previous = prev_date_by_date.get(visible_dates[date_index])
            if not previous:
                return None
            ref_row = add_ref_row(sheet_idx, previous, PREV_REF_TIME_20[row_type_idx])
            return f"FormulaRefs!{get_column_letter(close_col)}{ref_row}"

        if PREV_REF_TIME_20[row_type_idx] == "19":
            return f"{get_column_letter(close_col)}{block_start - 2}"
        return f"{get_column_letter(close_col)}{block_start - 1}"

    for idx in TIME_SHEET_IDXS:
        ws = out_wb.worksheets[idx]
        info = sheet_info[idx]
        style_ws = style_wb.worksheets[idx]
        for date_index, date_obj in enumerate(visible_dates):
            block_start = 3 + date_index * 6
            date_label = date_obj.strftime("%Y/%m/%d")
            for row_type_idx, time_str in enumerate(TIME_ROWS[idx]):
                row_num = block_start + row_type_idx
                _apply_row_style(ws, row_styles[idx], row_num, row_type_idx)
                ws.cell(row_num, 1).value = f"{date_label}-{time_str}时Close"

                for close_col in info["close_cols"]:
                    ws.cell(row_num, close_col).value = close_value(idx, date_obj, time_str, close_col)

                for close_col, change_col in info["direct_change_by_close"].items():
                    denominator = denominator_ref(idx, date_index, row_type_idx, close_col, block_start)
                    if denominator:
                        ws.cell(row_num, change_col).value = (
                            f"={get_column_letter(close_col)}{row_num}/{denominator}-1"
                        )
                    else:
                        ws.cell(row_num, change_col).value = None

                pattern_row = 9 + row_type_idx
                for col in info["derived_cols"]:
                    ws.cell(row_num, col).value = _same_row_formula(style_ws.cell(pattern_row, col).value, row_num)

    for idx in TIME_SHEET_IDXS:
        change_ws = out_wb.worksheets[CHANGE_SHEET_OFFSET + idx]
        time_ws = out_wb.worksheets[idx]
        info = sheet_info[idx]
        for date_index, date_obj in enumerate(visible_dates):
            block_start = 3 + date_index * 6
            date_label = date_obj.strftime("%Y/%m/%d")
            for row_type_idx in range(6):
                row_num = block_start + row_type_idx
                _apply_row_style(change_ws, row_styles[CHANGE_SHEET_OFFSET + idx], row_num, row_type_idx)
                change_ws.cell(row_num, 1).value = date_label if row_type_idx == 0 else None
                # Preserve the Time labels from the latest daily workbook.
                change_ws.cell(row_num, 2).value = style_wb.worksheets[CHANGE_SHEET_OFFSET + idx].cell(9 + row_type_idx, 2).value

                for col in range(3, change_ws.max_column + 1):
                    header = change_ws.cell(1, col).value
                    source_col = info["time_change_col_by_header"].get(header)
                    if source_col:
                        change_ws.cell(row_num, col).value = (
                            f"='{time_ws.title}'!{get_column_letter(source_col)}{row_num}"
                        )
                    else:
                        change_ws.cell(row_num, col).value = None

    if output_file is None:
        output_file = os.path.join(MONTHLY_REPORT_DIR, f"IG变化率_{month}_公式版.xlsx")

    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    out_wb.active = 0
    out_wb.save(output_path)
    style_wb.close()
    out_wb.close()
    return str(output_path)


def send_gmail_with_attachments(
    send_usr: str,
    send_pwd: str,
    receive_usr_list: List[str],
    attachment_paths: Iterable[str],
    email_title: str,
    content: str,
    email_server: str = "smtp.gmail.com",
    email_port: int = 587,
) -> bool:
    """Send one Gmail message with one or more attachments."""
    paths = [str(path) for path in attachment_paths]
    if not paths:
        print("No attachments configured; email not sent.")
        return False

    for path in paths:
        if not os.path.exists(path):
            print(f"Attachment does not exist: {path}")
            return False

    msg = MIMEMultipart()
    msg["Subject"] = email_title
    msg["From"] = send_usr
    msg["To"] = ", ".join(receive_usr_list)
    msg.attach(MIMEText(content, "plain", "utf-8"))

    for path in paths:
        with open(path, "rb") as file_obj:
            attachment = MIMEApplication(file_obj.read(), _subtype="xlsx")
        attachment.add_header(
            "Content-Disposition",
            "attachment",
            filename=os.path.basename(path),
        )
        msg.attach(attachment)
        print(f"Attached: {os.path.basename(path)}")

    try:
        smtp = SMTP(email_server, email_port, timeout=30)
        smtp.starttls()
        smtp.login(send_usr, send_pwd)
        smtp.sendmail(send_usr, receive_usr_list, msg.as_string())
        smtp.quit()
        print("Email sent successfully.")
        return True
    except Exception as exc:
        print(f"Email failed: {exc}")
        return False


def build_and_send_monthly_report(
    output_root_dir: str = OUTPUT_ROOT_DIR,
    report_dir: str = MONTHLY_REPORT_DIR,
    report_month: Optional[str] = None,
    send_email: bool = True,
) -> str:
    month = _parse_report_month(report_month)
    output_file = os.path.join(report_dir, f"IG变化率_{month}_公式版.xlsx")
    report_file = build_monthly_formula_report(
        output_root_dir=output_root_dir,
        output_file=output_file,
        report_month=month,
    )

    if send_email:
        from config import get_gmail_config

        gmail_config = get_gmail_config()
        sent = send_gmail_with_attachments(
            send_usr=gmail_config.send_usr,
            send_pwd=gmail_config.send_pwd,
            receive_usr_list=gmail_config.receive_usr_list,
            attachment_paths=[report_file],
            email_title=f"IG变化率累计公式版 - {month}",
            content="这是当天生成的IG变化率月度累计公式版，请查收。",
            email_server=gmail_config.email_server,
            email_port=gmail_config.email_port,
        )
        if not sent:
            raise RuntimeError("Monthly report was generated, but email sending failed")

    return report_file


def main() -> None:
    send_email_value = os.getenv("MONTHLY_REPORT_SEND_EMAIL", "true").strip().lower()
    send_email = send_email_value not in {"0", "false", "no", "off"}
    report_file = build_and_send_monthly_report(send_email=send_email)
    print(f"Monthly report ready: {report_file}")


if __name__ == "__main__":
    main()
