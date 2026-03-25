# -*- coding: utf-8 -*-
"""
一键跑完主程序（按你的原逻辑串联）：
Step 1) IG 抓取全量数据（1h + 30Min，全量不筛选不去重，标注Resolution，索引转伦敦时间）
Step 2) 更新模板日期（8个sheet：05/07/15/20时 + 05/07/15/20变化率）
Step 3) 从全量数据筛选出模板需要的时间点（1h: 05/07/15/19/20；30Min: 18:00/18:30）
Step 4) 把筛选后的 Close 写入模板的 05/07/15/20时 sheet
Step 5) （可选）Gmail 发送附件（支持多收件人）

说明：适配Python 3.8，移除zoneinfo依赖，仅使用pytz处理时区
"""

import os
import warnings
import logging
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from pandas import json_normalize

# ========== 依赖：IG抓取 ==========
from trading_ig import IGService
from trading_ig.rest import ApiExceededException
from tenacity import Retrying, wait_exponential, retry_if_exception_type

# ========== 依赖：模板处理 / 写入 ==========
from openpyxl import load_workbook

# ========== 时区（仅用pytz，兼容Python3.8） ==========
import pytz

# ========== 邮件（可选） ==========
from smtplib import SMTP
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication


# =============================================================================
# 0) 全局开关 & 路径配置（你按需改这里就行）
# =============================================================================

# 模板原文件（未改日期）
TEMPLATE_FILE = '/igcom/IG变化率表格(英区).xlsx'

# Step2 输出：日期已更新的模板
UPDATED_TEMPLATE_FILE = '/igcom/IG变化率表格_已更新.xlsx'

# Step4 输出：最终填好数据的表
FILLED_OUTPUT_FILE = '/igcom/IG变化率表格_已填好_05_07_15_20时.xlsx'

# Step3 输出：筛选后的历史数据Excel（会自动放到抓取输出目录里）
FILTERED_DATA_EXCEL_NAME = "All_Products_Full_1h_30min_filtered.xlsx"

# Step4 输入：你筛选后数据的“每产品sheet”的映射（保持你原逻辑）
PRODUCT_SHEET_MAP = {
    "US500": "US_500_Cash_(USD1)",
    "HK50": "Hong_Kong_HS50_Cash_(USD1)",
    "NIKKEI(Japan225)": "Japan_225_Cash_(USD1)",
    "USD/JPY(Yen)": "USD_JPY",
    "USD/SGD": "USD_SGD",
    "UK100(FTSE英国)": "UK_100_Cash_(USD1)",
    "GBP/USD": "GBP_USD",
    "France40(CAC法国)": "France_40_Cash_(USD1)",
    "EUR/USD": "EUR_USD",
    "USD/INR": "USD_INR_(USD1_Mini_Contract)",
    "Germany40": "Germany_40_Cash_(USD1)",
    "USD/CNH": "USD_CNH",
    "USD/TWD": "USD_TWD_(USD1_Mini_Contract)",
    "Australia200": "Australia_200_Cash_(USD1)",
    "AUD/USD": "AUD_USD",
    "USDKRW": "USD_KRW_(USD1_Mini_Contract)",
    "USDMXN": "USD_MXN",
}

TIME_SHEETS = ["05时", "07时", "15时", "20时"]
CHANGE_SHEETS = ["05变化率", "07变化率", "15变化率", "20变化率"]

# 你筛选规则（保持你原代码逻辑）
FILTER_1H_HOURS = {5, 7, 15, 19, 20}
FILTER_30MIN_TIMES = {"18:00", "18:30"}

# （可选）是否发送邮件
SEND_EMAIL = True

# ========== 多收件人列表（按需修改） ==========
# RECEIVE_USR_LIST = [
    
#     '18826728999@139.com'
# ]

RECEIVE_USR_LIST = [
    'shawncarpediem1121@gmail.com',
    'elitoc@163.com',
    '18826728999@139.com'
]

# =============================================================================
# 1) 日志与时区（仅用pytz，兼容Python3.8）
# =============================================================================

warnings.filterwarnings("ignore", category=FutureWarning)

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
console_handler.setFormatter(formatter)
logger.handlers = []
logger.addHandler(console_handler)

# 仅使用pytz定义时区（兼容Python3.8）
TZ_BEIJING = pytz.timezone("Asia/Shanghai")
TZ_LONDON = pytz.timezone("Europe/London")
TZ_UTC = pytz.UTC


# =============================================================================
# 2) Step 1 - IG抓取：你的原逻辑（全量1h+30Min，不过滤不去重，标注Resolution，索引转伦敦时间）
# =============================================================================

# IG账户配置（替换为实际信息）
class IGConfig:
    username = 'ZHONSH31795110'
    password = 'Zsy2713468YY'
    api_key = '2b4b1419858c86e1d3648b221921d27ab4c74962'
    acc_type = "LIVE"

EPIC_TO_NAME = {
    "IX.D.SPTRD.IFMM.IP": "US 500 Cash ($1)",
    "IX.D.HANGSENG.IFU.IP": "Hong Kong HS50 Cash ($1)",
    "IX.D.NIKKEI.IFM.IP": "Japan 225 Cash ($1)",
    "CS.D.USDJPY.CFD.IP": "USD/JPY",
    "CS.D.USDSGD.CFD.IP": "USD/SGD",
    "IX.D.FTSE.IFMM.IP": "UK 100 Cash ($1)",
    "CS.D.GBPUSD.CFD.IP": "GBP/USD",
    "IX.D.CAC.IFMM.IP": "France 40 Cash ($1)",
    "CS.D.EURUSD.CFD.IP": "EUR/USD",
    "CS.D.USDINR.MINI.IP": "USD/INR ($1 Mini Contract)",
    "IX.D.DAX.IFMS.IP": "Germany 40 Cash ($1)",
    "CS.D.USDCNH.CFD.IP": "USD/CNH",
    "CS.D.USDTWD.MINI.IP": "USD/TWD ($1 Mini Contract)",
    "IX.D.ASX.IFMM.IP": "Australia 200 Cash ($1)",
    "CS.D.AUDUSD.CFD.IP": "AUD/USD",
    "CS.D.USDKRW.MINI.IP": "USD/KRW ($1 Mini Contract)",
    "CS.D.USDMXN.CFD.IP": "USD/MXN",
}

# 30分钟粒度抓取范围（全天）
HALF_HOUR_RESOLUTION_RANGE = {"start_hour": 0, "end_hour": 23}


def safe_sheet_name(name: str) -> str:
    invalid_chars = r'\/:*?"<>|'
    for char in invalid_chars:
        name = name.replace(char, "_")
    return name[:31]


def safe_mid_prices(prices, version):
    """仅保留Close中间价，时间转换为英区时间（你的原逻辑）"""
    if len(prices) == 0:
        raise Exception("Historical price data not found")

    df = json_normalize(prices)

    if version == "3":
        df = df.set_index("snapshotTimeUTC")
        df = df.drop(columns=["snapshotTime"], errors="ignore")
        df.index = pd.to_datetime(df.index, format="ISO8601")
    else:
        df = df.set_index("snapshotTime")
        from trading_ig.utils import DATE_FORMATS

        date_format = DATE_FORMATS[int(version)]
        df.index = pd.to_datetime(df.index, format=date_format)

    # UTC转英区时间（仅用pytz，兼容Python3.8）
    df.index = df.index.tz_localize(TZ_UTC).tz_convert(TZ_LONDON).tz_localize(None)
    df.index.name = "DateTime (London)"

    df["Close"] = df[["closePrice.bid", "closePrice.ask"]].mean(axis=1)

    drop_cols = [
        "openPrice.lastTraded",
        "closePrice.lastTraded",
        "highPrice.lastTraded",
        "lowPrice.lastTraded",
        "openPrice.bid",
        "openPrice.ask",
        "closePrice.bid",
        "closePrice.ask",
        "highPrice.bid",
        "highPrice.ask",
        "lowPrice.bid",
        "lowPrice.ask",
        "lastTradedVolume",
    ]
    df = df.drop(columns=[c for c in drop_cols if c in df.columns], errors="ignore")
    return df


def fetch_data_by_resolution(ig_service, epic: str, resolution: str, start_date_str: str, end_date_str: str) -> pd.DataFrame:
    """按指定粒度抓取单个产品数据，并标注粒度（你的原逻辑）"""
    try:
        response = ig_service.fetch_historical_prices_by_epic(
            epic=epic,
            resolution=resolution,
            start_date=start_date_str,
            end_date=end_date_str,
            format=safe_mid_prices,
        )
        df = response["prices"]
        df["Resolution"] = resolution
        return df
    except Exception as e:
        logger.warning(f"Failed to fetch {resolution} data for {epic}: {str(e)}")
        return pd.DataFrame()


def get_multiple_historical_prices_full(
    epic_list,
    start_date=None,
    end_date=None,
    days=2,  # 抓取北京时间2天内数据（不重置时分秒）
    save_individual=True,
    save_combined=True,
):
    """
    原逻辑保留，修改点：
    1. start_date = 北京时间end_date - N天（不重置为00:00）
    2. 直接传北京时间字符串给IG API（ISO8601格式）
    3. 全量1h+30Min、不过滤、标注Resolution、索引转伦敦时间不变
    """
    retryer = Retrying(wait=wait_exponential(), retry=retry_if_exception_type(ApiExceededException))
    ig_service = IGService(
        IGConfig.username,
        IGConfig.password,
        IGConfig.api_key,
        IGConfig.acc_type,
        retryer=retryer,
        use_rate_limiter=True,
    )

    all_data = {}
    combined_data = []

    # ========== 基于北京时间计算，不重置时间 ==========
    if end_date is None:
        end_date = datetime.now(TZ_BEIJING)  # 当前北京时间（带时区）
    else:
        if not end_date.tzinfo:
            end_date = TZ_BEIJING.localize(end_date)  # 补全时区
        else:
            end_date = end_date.astimezone(TZ_BEIJING)

    if start_date is None:
        start_date = end_date - timedelta(days=2)  # 北京时间减N天，不重置时分秒
    else:
        if not start_date.tzinfo:
            start_date = TZ_BEIJING.localize(start_date)
        else:
            start_date = start_date.astimezone(TZ_BEIJING)

    # ========== 直接传北京时间字符串给API（ISO8601格式） ==========
    start_date_str = start_date.strftime("%Y-%m-%dT%H:%M:%S")
    end_date_str = end_date.strftime("%Y-%m-%dT%H:%M:%S")

    # 北京时间转伦敦时间/UTC（仅用于日志打印）
    start_date_london = start_date.astimezone(TZ_LONDON)
    end_date_london = end_date.astimezone(TZ_LONDON)
    start_date_utc = start_date.astimezone(TZ_UTC)
    end_date_utc = end_date.astimezone(TZ_UTC)

    print("======================================")
    print("✅ Step 1/5: IG 抓取全量数据（1h + 30Min，保留全量，标注Resolution，伦敦时间索引）")
    print(f"📅 抓取范围（北京时间）: {start_date.strftime('%Y-%m-%d %H:%M:%S')} 至 {end_date.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"📅 Date Range (London): {start_date_london.strftime('%Y-%m-%d %H:%M:%S')} to {end_date_london.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"📅 Date Range (UTC)   : {start_date_utc.strftime('%Y-%m-%d %H:%M:%S')} to {end_date_utc.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"📤 API传入时间（北京时间）: start={start_date_str}, end={end_date_str}")
    print(f"📊 Total Products: {len(epic_list)}")
    print("======================================\n")

    output_dir_abs = None
    combined_excel_path_abs = None

    try:
        ig_service.create_session()
        print("✅ IG Session Created Successfully")

        output_dir = f"historical_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        os.makedirs(output_dir, exist_ok=True)
        output_dir_abs = os.path.abspath(output_dir)
        print(f"📁 Output Directory (Step1 outputs): {output_dir_abs}")

        for i, epic in enumerate(epic_list, 1):
            product_name = EPIC_TO_NAME.get(epic, epic)
            print(f"\n--- Fetching {i}/{len(epic_list)}: {product_name} ({epic}) ---")

            print("🔹 Fetching 1h resolution data...")
            df_1h = fetch_data_by_resolution(ig_service, epic, "1h", start_date_str, end_date_str)
            if df_1h.empty:
                print(f"⚠️ No 1h data for {product_name}")
            else:
                print(f"✅ 1h data fetched: {len(df_1h)} records")

            print("🔹 Fetching 30Min resolution data...")
            df_30min = fetch_data_by_resolution(ig_service, epic, "30Min", start_date_str, end_date_str)
            if not df_30min.empty:
                df_30min = df_30min[
                    (df_30min.index.hour >= HALF_HOUR_RESOLUTION_RANGE["start_hour"])
                    & (df_30min.index.hour <= HALF_HOUR_RESOLUTION_RANGE["end_hour"])
                ]
                print(f"✅ 30Min data fetched: {len(df_30min)} records")
            else:
                print("ℹ️ No 30Min data fetched")

            df_list = []
            if not df_1h.empty:
                df_list.append(df_1h)
            if not df_30min.empty:
                df_list.append(df_30min)

            if not df_list:
                print(f"⚠️ No valid data for {product_name}, skip")
                continue

            df_combined = pd.concat(df_list).sort_index()
            print(f"✅ Final merged data: {len(df_combined)} records (1h + 30Min full)")

            df_final = df_combined.copy()
            df_final["Product Name"] = product_name
            df_final["Epic"] = epic

            all_data[epic] = df_final
            combined_data.append(df_final)

            if save_individual:
                safe_filename = safe_sheet_name(product_name).replace(" ", "_").replace("$", "USD")
                csv_path = os.path.join(output_dir, f"{safe_filename}_full_1h_30min_{datetime.now().strftime('%Y%m%d')}.csv")
                df_final.to_csv(csv_path, encoding="utf-8")
                print(f"💾 Saved CSV: {os.path.abspath(csv_path)}")

        if save_combined and combined_data:
            df_combined_all = pd.concat(combined_data, ignore_index=False).sort_index()

            combined_csv = os.path.join(output_dir, f"All_Products_Full_1h_30min_{datetime.now().strftime('%Y%m%d')}.csv")
            df_combined_all.to_csv(combined_csv, encoding="utf-8")
            print(f"\n💾 All Products Combined CSV: {os.path.abspath(combined_csv)}")

            combined_excel = os.path.join(output_dir, f"All_Products_Full_1h_30min_{datetime.now().strftime('%Y%m%d')}.xlsx")
            with pd.ExcelWriter(combined_excel, engine="openpyxl") as writer:
                df_combined_all.to_excel(writer, sheet_name="All_Full_Data", index=True)
                for epic, df in all_data.items():
                    sheet_name = safe_sheet_name(EPIC_TO_NAME.get(epic, epic))
                    df.to_excel(writer, sheet_name=sheet_name, index=True)

            combined_excel_path_abs = os.path.abspath(combined_excel)
            print(f"💾 All Products Combined Excel: {combined_excel_path_abs}")

        return all_data, output_dir_abs, combined_excel_path_abs

    finally:
        try:
            ig_service.logout()
        except Exception:
            pass
        print("\n🔚 Session Closed\n")


# =============================================================================
# 3) Step 2 - 更新模板日期（8个sheet）
# =============================================================================

def update_template_dates_uk(TARGET_FILE: str, OUTPUT_FILE: str):
    """更新模板日期（基于英国时间）"""
    uk_now = datetime.now(TZ_LONDON)
    today_str = uk_now.strftime("%Y/%m/%d")
    yesterday_str = (uk_now - timedelta(days=1)).strftime("%Y/%m/%d")

    print("======================================")
    print("✅ Step 2/5: 更新模板日期（英国时间 Europe/London）")
    print(f"🧾 Read Template : {TARGET_FILE}")
    print(f"🧾 Write Updated : {OUTPUT_FILE}")
    print("📌 当前英国时间:", uk_now.strftime("%Y-%m-%d %H:%M:%S %Z"))
    print("📅 Yesterday (UK):", yesterday_str)
    print("📅 Today     (UK):", today_str)
    print("======================================\n")

    if not os.path.exists(TARGET_FILE):
        raise FileNotFoundError(f"❌ 找不到模板文件：{TARGET_FILE}")

    wb = load_workbook(TARGET_FILE)

    def update_time_sheet_cell(cell_value, new_date_str):
        if not isinstance(cell_value, str):
            return None
        if "-" not in cell_value or "时Close" not in cell_value:
            return None
        parts = cell_value.split("-", 1)
        suffix = parts[1]
        return f"{new_date_str}-{suffix}"

    def normalize_change_sheet_date_cell(cell_value):
        if cell_value is None:
            return None
        if isinstance(cell_value, datetime):
            return cell_value.strftime("%Y/%m/%d")
        if isinstance(cell_value, str):
            s = cell_value.strip()
            if s.count("/") == 2 and len(s) >= 10:
                return s[:10]
            if len(s) >= 10 and s[4] == "-" and s[7] == "-":
                return s[:10].replace("-", "/")
        return None

    # 更新时间sheet（A3-A8=昨天, A9-A14=今天）
    print("✅ Step 2.1: 更新 time_sheets (A3-A14)")
    for sheet_name in TIME_SHEETS:
        print(f"\n🧾 处理 sheet: {sheet_name}")
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠ 不存在，跳过：{sheet_name}")
            continue
        ws = wb[sheet_name]

        for row in range(3, 9):
            cell = ws.cell(row=row, column=1)
            old = cell.value
            new = update_time_sheet_cell(old, yesterday_str)
            print(f"  - {sheet_name}!A{row} 原值: {old}")
            if new is not None and new != old:
                cell.value = new
                print(f"    ✅ 更新为: {new}")
            else:
                print("    ↪ 跳过")

        for row in range(9, 15):
            cell = ws.cell(row=row, column=1)
            old = cell.value
            new = update_time_sheet_cell(old, today_str)
            print(f"  - {sheet_name}!A{row} 原值: {old}")
            if new is not None and new != old:
                cell.value = new
                print(f"    ✅ 更新为: {new}")
            else:
                print("    ↪ 跳过")

    # 更新变化率sheet（A3=昨天, A9=今天）
    print("\n✅ Step 2.2: 更新 change_sheets (A3/A9)")
    for sheet_name in CHANGE_SHEETS:
        print(f"\n🧾 处理 sheet: {sheet_name}")
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠ 不存在，跳过：{sheet_name}")
            continue
        ws = wb[sheet_name]

        cell = ws.cell(row=3, column=1)
        old = cell.value
        norm = normalize_change_sheet_date_cell(old)
        print(f"  - {sheet_name}!A3 原值: {old} | 识别为: {norm}")
        if norm is not None:
            cell.value = yesterday_str
            print(f"    ✅ 更新为: {yesterday_str}")
        else:
            print("    ⚠ 无法识别，跳过 A3")

        cell = ws.cell(row=9, column=1)
        old = cell.value
        norm = normalize_change_sheet_date_cell(old)
        print(f"  - {sheet_name}!A9 原值: {old} | 识别为: {norm}")
        if norm is not None:
            cell.value = today_str
            print(f"    ✅ 更新为: {today_str}")
        else:
            print("    ⚠ 无法识别，跳过 A9")

    wb.save(OUTPUT_FILE)
    print(f"\n🎉 Step 2 完成：已保存 -> {OUTPUT_FILE}\n")
    return yesterday_str, today_str


# =============================================================================
# 4) Step 3 - 对全量Excel进行筛选（按你的统一规则）
# =============================================================================

def filter_historical_data_full_to_template_times(input_excel_path: str, output_excel_path: str):
    """从全量数据筛选模板需要的时间点"""
    print("======================================")
    print("✅ Step 3/5: 从全量历史数据筛选出模板需要的时间点")
    print(f"🧾 Read  Full Excel : {input_excel_path}")
    print(f"🧾 Write Filtered   : {output_excel_path}")
    print(f"🔧 Rule 1h hours     : {sorted(list(FILTER_1H_HOURS))}")
    print(f"🔧 Rule 30Min times  : {sorted(list(FILTER_30MIN_TIMES))}")
    print("======================================\n")

    df = pd.read_excel(
        input_excel_path,
        sheet_name="All_Full_Data",
        index_col="DateTime (London)",
        parse_dates=["DateTime (London)"],
    )
    print(f"✅ 成功读取全量数据：{len(df)} 条记录")

    filtered_data = []
    for product_name, group in df.groupby("Product Name"):
        print(f"\n🔍 筛选产品：{product_name}")
        group_1h = group[group["Resolution"] == "1h"].copy()
        group_30min = group[group["Resolution"] == "30Min"].copy()

        # 筛选1h数据（指定小时）
        group_1h_filtered = group_1h[group_1h.index.hour.isin(FILTER_1H_HOURS)]

        # 筛选30Min数据（指定时间）
        group_30min["time_str"] = group_30min.index.strftime("%H:%M")
        group_30min_filtered = group_30min[group_30min["time_str"].isin(FILTER_30MIN_TIMES)]
        group_30min_filtered = group_30min_filtered.drop(columns=["time_str"], errors="ignore")

        # 合并筛选结果
        group_filtered = pd.concat([group_1h_filtered, group_30min_filtered]).sort_index()
        print(
            f"   1h筛选后：{len(group_1h_filtered)} 条 | "
            f"30Min筛选后：{len(group_30min_filtered)} 条 | "
            f"总计：{len(group_filtered)} 条"
        )
        filtered_data.append(group_filtered)

    df_filtered = pd.concat(filtered_data).sort_index()
    print(f"\n📊 最终筛选结果：{len(df_filtered)} 条记录（原始：{len(df)} 条）")

    # 保存筛选后数据
    Path(os.path.dirname(output_excel_path)).mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_excel_path, engine="openpyxl") as writer:
        df_filtered.to_excel(writer, sheet_name="Filtered_Full_Data", index=True)
        for product_name, group in df_filtered.groupby("Product Name"):
            safe_name = product_name.replace("/", "_").replace("$", "USD").replace(" ", "_")[:31]
            group.to_excel(writer, sheet_name=safe_name, index=True)

    print(f"💾 Step 3 完成：筛选后数据已保存 -> {output_excel_path}\n")
    return output_excel_path


# =============================================================================
# 5) Step 4 - 把筛选后的 Close 写入更新过日期的模板（05/07/15/20时）
# =============================================================================

def parse_timestamp_label(label: str) -> datetime:
    """解析模板中的时间标签为datetime对象"""
    label = str(label).strip()
    date_part, time_part = label.split("-")
    date_obj = datetime.strptime(date_part.strip(), "%Y/%m/%d")

    time_str = time_part.split("时")[0]
    if ":" in time_str:
        hour_str, minute_str = time_str.split(":")
        hour = int(hour_str)
        minute = int(minute_str)
    else:
        hour = int(time_str)
        minute = 0

    return date_obj.replace(hour=hour, minute=minute, second=0, microsecond=0)


def fill_template_with_close_data(source_file: str, template_file: str, output_file: str):
    """将筛选后的Close数据写入模板"""
    print("======================================")
    print("✅ Step 4/5: 把筛选后的 Close 数据写入模板（05/07/15/20时）")
    print(f"🧾 Read  Filtered Data Excel : {source_file}")
    print(f"🧾 Read  Updated Template    : {template_file}")
    print(f"🧾 Write Filled Output       : {output_file}")
    print("======================================\n")

    if not os.path.exists(source_file):
        raise FileNotFoundError(f"源数据文件不存在：{source_file}")
    if not os.path.exists(template_file):
        raise FileNotFoundError(f"模板文件不存在：{template_file}")

    # 读取源数据
    print("📥 正在读取源数据工作簿……")
    product_data = {}
    source_sheets = pd.ExcelFile(source_file).sheet_names
    print(f"🔍 源文件包含工作表：{source_sheets}")

    for header_name, sheet_name in PRODUCT_SHEET_MAP.items():
        print(f"  - 尝试读取 {header_name} <- sheet: {sheet_name}")
        try:
            df = pd.read_excel(source_file, sheet_name=sheet_name)
        except ValueError as e:
            print(f"  ❌ 无法找到工作表 '{sheet_name}'，错误：{e}")
            continue

        if "DateTime (London)" not in df.columns:
            print(f"  ❌ 工作表 '{sheet_name}' 缺少 'DateTime (London)' 列，跳过")
            continue
        df["DateTime (London)"] = pd.to_datetime(df["DateTime (London)"])
        df = df.set_index("DateTime (London)")

        if "Close" not in df.columns:
            print(f"  ❌ 工作表 '{sheet_name}' 缺少 'Close' 列，跳过")
            continue

        keep_cols = ["Close"]
        if "Resolution" in df.columns:
            keep_cols.append("Resolution")
        product_data[header_name] = df[keep_cols]

    print(f"\n✅ 源数据读取完成，共加载 {len(product_data)} 个产品")

    # 写入模板
    print("📗 正在打开目标模板工作簿……")
    wb = load_workbook(template_file)

    for sheet_name in TIME_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"⚠ 警告：模板中找不到工作表 '{sheet_name}'，跳过。")
            continue

        ws = wb[sheet_name]
        print(f"\n📝 处理工作表：{sheet_name}")

        max_row = ws.max_row
        max_col = ws.max_column

        # 映射列到产品
        col_to_product = {}
        for col in range(2, max_col + 1):
            header = ws.cell(row=1, column=col).value
            if header and header in PRODUCT_SHEET_MAP:
                col_to_product[col] = header
                print(f"  - 第 {col} 列 对应产品：{header}")

        if not col_to_product:
            print("  ⚠ 第1行没有找到任何已配置的产品列，跳过该表。")
            continue

        # 逐行写入数据
        for row in range(3, max_row + 1):
            label = ws.cell(row=row, column=1).value
            if not label:
                continue

            try:
                ts = parse_timestamp_label(label)
            except Exception as e:
                print(f"  ⚠ 第 {row} 行 A 列无法解析时间：{label} -> {e}")
                continue

            for col, product_header in col_to_product.items():
                df = product_data.get(product_header)
                if df is None:
                    continue

                if ts in df.index:
                    close_value = df.loc[ts, "Close"]
                    if hasattr(close_value, "iloc"):
                        close_value = close_value.iloc[0]
                    close_value = float(close_value)

                    # 18:30用18:00数据
                    if ts.hour == 18 and ts.minute == 30:
                        ts_1800 = ts.replace(minute=0)
                        if ts_1800 in df.index:
                            close_value = df.loc[ts_1800, "Close"]
                            print(f"    调整：18:30 数据使用 18:00 的 Close 值：{close_value}")

                    # 18:00用18:30数据
                    if ts.hour == 18 and ts.minute == 0:
                        ts_1830 = ts.replace(minute=30)
                        if ts_1830 in df.index:
                            close_value = df.loc[ts_1830, "Close"]
                            print(f"    调整：18:00 数据使用 18:30 的 Close 值：{close_value}")

                    ws.cell(row=row, column=col).value = close_value
                    print(
                        f"    写入 {sheet_name}!{ws.cell(row=row, column=col).coordinate} "
                        f"<- {product_header} {ts} Close={close_value}"
                    )

                else:
                    # 20:00用19:00补数据
                    if ts.hour == 20 and ts.minute == 0:
                        ts_1900 = ts.replace(hour=19, minute=0)
                        if ts_1900 in df.index:
                            close_value = df.loc[ts_1900, "Close"]
                            if hasattr(close_value, "iloc"):
                                close_value = close_value.iloc[0]
                            close_value = float(close_value)

                            ws.cell(row=row, column=col).value = close_value
                            print(
                                f"    调整：20:00 数据使用 19:00 的 Close 值：{close_value} "
                                f"已写入 {sheet_name}!{ws.cell(row=row, column=col).coordinate}"
                            )
                        else:
                            print(f"    ⚠ {product_header} 20:00 数据缺失，且19:00也没有数据，不写入。")
                    else:
                        print(f"    ⚠ {product_header} 缺少 {ts} 的数据，不写入。")

    wb.save(output_file)
    print(f"\n🎉 Step 4 完成：已保存为 -> {output_file}\n")
    return output_file


# =============================================================================
# 6) Step 5 - （可选）Gmail发送（支持多收件人）
# =============================================================================

def send_gmail_with_attachment(send_usr, send_pwd, receive_usr_list, attachment_path, email_title, content):
    """
    支持多收件人的Gmail邮件发送函数
    :param send_usr: 发件人邮箱
    :param send_pwd: 发件人授权码
    :param receive_usr_list: 收件人列表
    :param attachment_path: 附件路径
    :param email_title: 邮件标题
    :param content: 邮件正文
    """
    print("======================================")
    print("✅ Step 5/5: 发送 Gmail 邮件（含附件）")
    print(f"📨 From: {send_usr}")
    print(f"📨 To  : {', '.join(receive_usr_list)}")
    print(f"📎 Attachment: {attachment_path}")
    print("======================================\n")

    email_server = "smtp.gmail.com"
    email_port = 587

    # 构建邮件
    msg = MIMEMultipart()
    msg["Subject"] = email_title
    msg["From"] = send_usr
    msg["To"] = ", ".join(receive_usr_list)  # 多收件人用逗号分隔
    msg.attach(MIMEText(content, "plain", "utf-8"))

    # 添加附件
    if os.path.exists(attachment_path):
        with open(attachment_path, "rb") as f:
            attachment = MIMEApplication(f.read(), _subtype="xlsx")
            attachment.add_header("Content-Disposition", "attachment", filename=os.path.basename(attachment_path))
            msg.attach(attachment)
        print(f"✅ 已添加附件：{os.path.basename(attachment_path)}")
    else:
        print(f"❌ 附件不存在：{attachment_path}")
        return

    # 发送邮件
    try:
        smtp = SMTP(email_server, email_port, timeout=30)
        smtp.starttls()
        smtp.login(send_usr, send_pwd)
        smtp.sendmail(send_usr, receive_usr_list, msg.as_string())  # 传入收件人列表
        smtp.quit()
        print("✅ Gmail邮件（含附件）发送成功！\n")
    except Exception as e:
        print(f"❌ 发送失败：{str(e)}\n")


# =============================================================================
# 7) 主程序：一键跑完
# =============================================================================

def main():
    print("\n" + "=" * 70)
    print("🚀 一键跑完主程序启动")
    print("=" * 70)

    # Step 1: 抓取全量数据
    epic_list = [
        "IX.D.SPTRD.IFMM.IP",
        "IX.D.HANGSENG.IFU.IP",
        "IX.D.NIKKEI.IFM.IP",
        "CS.D.USDJPY.CFD.IP",
        "CS.D.USDSGD.CFD.IP",
        "IX.D.FTSE.IFMM.IP",
        "CS.D.GBPUSD.CFD.IP",
        "IX.D.CAC.IFMM.IP",
        "CS.D.EURUSD.CFD.IP",
        "CS.D.USDINR.MINI.IP",
        "IX.D.DAX.IFMS.IP",
        "CS.D.USDCNH.CFD.IP",
        "CS.D.USDTWD.MINI.IP",
        "IX.D.ASX.IFMM.IP",
        "CS.D.AUDUSD.CFD.IP",
        "CS.D.USDKRW.MINI.IP",
        "CS.D.USDMXN.CFD.IP",
    ]

    

    days = 2  # 抓取近2天（48小时）数据
    all_data, output_dir_abs, full_excel_abs = get_multiple_historical_prices_full(
        epic_list=epic_list,
        start_date=None,
        end_date=None,
        days=days,
        save_individual=True,
        save_combined=True,
    )

    if not full_excel_abs:
        raise RuntimeError("❌ Step1 未生成合并Excel，无法继续。")
    print(f"\n✅ Step1 产物确认：")
    print(f"📁 抓取输出目录: {output_dir_abs}")
    print(f"📄 全量合并Excel : {full_excel_abs}\n")

    # Step 2: 更新模板日期
    update_template_dates_uk(TEMPLATE_FILE, UPDATED_TEMPLATE_FILE)

    # Step 3: 筛选数据
    filtered_excel_abs = os.path.join(output_dir_abs, FILTERED_DATA_EXCEL_NAME)
    filter_historical_data_full_to_template_times(full_excel_abs, filtered_excel_abs)

    # Step 4: 填充模板
    fill_template_with_close_data(
        source_file=filtered_excel_abs,
        template_file=UPDATED_TEMPLATE_FILE,
        output_file=FILLED_OUTPUT_FILE,
    )

    # Step 5: 发送邮件（可选）
    if SEND_EMAIL:
        send_usr = 'xieminggen@gmail.com'  # 你的发件人邮箱
        send_pwd = 'hqsivzuksymeyuga'      # 你的Gmail授权码
        email_title = f"Excel数据附件 - {datetime.now().strftime('%Y%m%d')} - 变化率表格"
        content = "这是用Python脚本发送的邮件(变化率表格)，附带Excel数据附件，请查收！"

        send_gmail_with_attachment(
            send_usr=send_usr,
            send_pwd=send_pwd,
            receive_usr_list=RECEIVE_USR_LIST,  # 多收件人列表
            attachment_path=FILLED_OUTPUT_FILE,
            email_title=email_title,
            content=content,
        )
    else:
        print("ℹ️ Step5 已跳过（SEND_EMAIL = False）\n")

    # 输出完成信息
    print("=" * 70)
    print("🎉 全流程完成！")
    print(f"📁 抓取数据输出目录: {output_dir_abs}")
    print(f"📄 全量数据Excel    : {full_excel_abs}")
    print(f"📄 筛选后数据Excel  : {filtered_excel_abs}")
    print(f"📄 更新后模板       : {UPDATED_TEMPLATE_FILE}")
    print(f"📄 最终填好数据表   : {FILLED_OUTPUT_FILE}")
    print("=" * 70 + "\n")


if __name__ == "__main__":
    main()请你分析一下，这个处理的文件是什么
