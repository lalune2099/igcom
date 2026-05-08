# -*- coding: utf-8 -*-
"""
改动目标（已实现）：
- 每天运行一次，只抓取“昨天（伦敦时间）”的数据（00:00-24:00）
- 抓到的数据追加到一个“累计总表 Excel”里，避免重复抓取导致API额度不足
- 后续更新模板/筛选/填充/发邮件逻辑保持你的原方式

Step 1) IG 增量抓取（仅昨天 London day，1h + 30Min，全量不去重，标注Resolution，索引为 London 无时区）
Step 2) 更新模板日期（8个sheet：05/07/15/20时 + 05/07/15/20变化率）
Step 3) 从“累计总表”筛选出模板需要的时间点
Step 4) 把筛选后的 Close 写入模板的 05/07/15/20时 sheet
Step 5) （可选）Gmail 发送附件（支持多收件人）
"""

import os
import warnings
import logging
from datetime import datetime, timedelta, time
from pathlib import Path

import pandas as pd
from pandas import json_normalize

from trading_ig import IGService
from trading_ig.rest import ApiExceededException
from tenacity import Retrying, wait_exponential, retry_if_exception_type

from openpyxl import load_workbook

import pytz

from smtplib import SMTP
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from config import get_gmail_config, get_ig_account


# =============================================================================
# 0) 全局路径配置
# =============================================================================

# 模板原文件（未改日期）
TEMPLATE_FILE = '/igcom/IG变化率表格(英区).xlsx'
# Step2 输出：日期已更新的模板
UPDATED_TEMPLATE_FILE = '/igcom/IG变化率表格_已更新.xlsx'
# Step4 输出：最终填好数据的表
FILLED_OUTPUT_FILE = '/igcom/IG变化率表格_已填好_05_07_15_20时.xlsx'

# ✅ 新增：累计总表（核心）
ACCUMULATED_EXCEL_FILE = '/igcom/ig_accumulated_full_1h_30min.xlsx'

# Step3 输出：筛选后的历史数据Excel（会自动放到输出目录里）
FILTERED_DATA_EXCEL_NAME = "All_Products_Full_1h_30min_filtered.xlsx"

# Step4 输入：你筛选后数据的“每产品sheet”的映射（保持你原逻辑）
PRODUCT_SHEET_MAP = {
    "US500": "US_500_Cash_(USD1)",
    "HK50": "Hong_Kong_HS50_Cash_(USD1)",
    "NIKKEI(Japan225)": "Japan_225_Cash_(USD1)",
    "USD/JPY(Yen)": "USD_JPY",
    "USD/SGD": "USD_SGD",
    "UK100(FTSE英国)": "UK_100_Cash_(USD1)",
    "GBP/USD": "GBP_USD",
    "France40(CAC法国)": "France_40_Cash_(USD1)",
    "EUR/USD": "EUR_USD",
    "USD/INR": "USD_INR_(USD1_Mini_Contract)",
    "Germany40": "Germany_40_Cash_(USD1)",
    "USD/CNH": "USD_CNH",
    "USD/TWD": "USD_TWD_(USD1_Mini_Contract)",
    "Australia200": "Australia_200_Cash_(USD1)",
    "AUD/USD": "AUD_USD",
    "USDKRW": "USD_KRW_(USD1_Mini_Contract)",
    "USDMXN": "USD_MXN",
}

TIME_SHEETS = ["05时", "07时", "15时", "20时"]
CHANGE_SHEETS = ["05变化率", "07变化率", "15变化率", "20变化率"]

# 你筛选规则（保持你原代码逻辑）
FILTER_1H_HOURS = {5, 7, 15, 19, 20}
FILTER_30MIN_TIMES = {"18:00", "18:30"}

# （可选）是否发送邮件
SEND_EMAIL = True

# Gmail recipients are loaded from GMAIL_RECIPIENTS when SEND_EMAIL is enabled.


# =============================================================================
# 1) 日志与时区
# =============================================================================

warnings.filterwarnings("ignore", category=FutureWarning)

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
console_handler.setFormatter(formatter)
logger.handlers = []
logger.addHandler(console_handler)

TZ_BEIJING = pytz.timezone("Asia/Shanghai")
TZ_LONDON = pytz.timezone("Europe/London")
TZ_UTC = pytz.UTC


# =============================================================================
# 2) Step 1 - IG 增量抓取：只抓“昨天 London day”
# =============================================================================




# IG账户配置：默认使用当前脚本对应的账号，可用 IG_PROFILE 覆盖。
DEFAULT_IG_PROFILE = "ACCOUNT1"
_ig_account = get_ig_account(DEFAULT_IG_PROFILE)


class IGConfig:
    username = _ig_account.username
    password = _ig_account.password
    api_key = _ig_account.api_key
    acc_type = _ig_account.acc_type
EPIC_TO_NAME = {
    "IX.D.SPTRD.IFMM.IP": "US 500 Cash ($1)",
    "IX.D.HANGSENG.IFU.IP": "Hong Kong HS50 Cash ($1)",
    "IX.D.NIKKEI.IFM.IP": "Japan 225 Cash ($1)",
    "CS.D.USDJPY.CFD.IP": "USD/JPY",
    "CS.D.USDSGD.CFD.IP": "USD/SGD",
    "IX.D.FTSE.IFMM.IP": "UK 100 Cash ($1)",
    "CS.D.GBPUSD.CFD.IP": "GBP/USD",
    "IX.D.CAC.IFMM.IP": "France 40 Cash ($1)",
    "CS.D.EURUSD.CFD.IP": "EUR/USD",
    "CS.D.USDINR.MINI.IP": "USD/INR ($1 Mini Contract)",
    "IX.D.DAX.IFMS.IP": "Germany 40 Cash ($1)",
    "CS.D.USDCNH.CFD.IP": "USD/CNH",
    "CS.D.USDTWD.MINI.IP": "USD/TWD ($1 Mini Contract)",
    "IX.D.ASX.IFMM.IP": "Australia 200 Cash ($1)",
    "CS.D.AUDUSD.CFD.IP": "AUD/USD",
    "CS.D.USDKRW.MINI.IP": "USD/KRW ($1 Mini Contract)",
    "CS.D.USDMXN.CFD.IP": "USD/MXN",
}

HALF_HOUR_RESOLUTION_RANGE = {"start_hour": 0, "end_hour": 23}


def safe_sheet_name(name: str) -> str:
    invalid_chars = r'\/:*?"<>|'
    for char in invalid_chars:
        name = name.replace(char, "_")
    return name[:31]


def safe_mid_prices(prices, version):
    """仅保留Close中间价，时间转换为伦敦时间（无时区）"""
    if len(prices) == 0:
        raise Exception("Historical price data not found")

    df = json_normalize(prices)

    if version == "3":
        df = df.set_index("snapshotTimeUTC")
        df = df.drop(columns=["snapshotTime"], errors="ignore")
        df.index = pd.to_datetime(df.index, format="ISO8601")
    else:
        df = df.set_index("snapshotTime")
        from trading_ig.utils import DATE_FORMATS
        date_format = DATE_FORMATS[int(version)]
        df.index = pd.to_datetime(df.index, format=date_format)

    # UTC -> London -> drop tz
    df.index = df.index.tz_localize(TZ_UTC).tz_convert(TZ_LONDON).tz_localize(None)
    df.index.name = "DateTime (London)"

    df["Close"] = df[["closePrice.bid", "closePrice.ask"]].mean(axis=1)

    drop_cols = [
        "openPrice.lastTraded", "closePrice.lastTraded", "highPrice.lastTraded", "lowPrice.lastTraded",
        "openPrice.bid", "openPrice.ask", "closePrice.bid", "closePrice.ask",
        "highPrice.bid", "highPrice.ask", "lowPrice.bid", "lowPrice.ask",
        "lastTradedVolume",
    ]
    df = df.drop(columns=[c for c in drop_cols if c in df.columns], errors="ignore")
    return df


def fetch_data_by_resolution(ig_service, epic: str, resolution: str, start_date_str: str, end_date_str: str) -> pd.DataFrame:
    """按指定粒度抓取单个产品数据，并标注粒度"""
    try:
        response = ig_service.fetch_historical_prices_by_epic(
            epic=epic,
            resolution=resolution,
            start_date=start_date_str,
            end_date=end_date_str,
            format=safe_mid_prices,
        )
        df = response["prices"]
        df["Resolution"] = resolution
        return df
    except Exception as e:
        logger.warning(f"Failed to fetch {resolution} data for {epic}: {str(e)}")
        return pd.DataFrame()


def london_yesterday_range_to_beijing_api_strings():
    """
    计算“昨天（伦敦日历日）00:00 到 今天（伦敦）00:00”
    然后转换为北京时间，用作API入参字符串（不带时区，按你原来API方式）
    """
    now_london = datetime.now(TZ_LONDON)

    today_london_date = now_london.date()
    yesterday_london_date = today_london_date - timedelta(days=1)

    start_london = TZ_LONDON.localize(datetime.combine(yesterday_london_date, time(0, 0, 0)))
    end_london = TZ_LONDON.localize(datetime.combine(today_london_date, time(0, 0, 0)))

    start_bj = start_london.astimezone(TZ_BEIJING)
    end_bj = end_london.astimezone(TZ_BEIJING)

    start_str = start_bj.strftime("%Y-%m-%dT%H:%M:%S")
    end_str = end_bj.strftime("%Y-%m-%dT%H:%M:%S")

    return start_london, end_london, start_bj, end_bj, start_str, end_str


def load_accumulated_excel(accumulated_path: str):
    """读取累计总表（如果不存在返回None）"""
    if not os.path.exists(accumulated_path):
        return None

    df_all = pd.read_excel(
        accumulated_path,
        sheet_name="All_Full_Data",
        index_col="DateTime (London)",
        parse_dates=["DateTime (London)"],
    )
    return df_all


def save_accumulated_excel(accumulated_path: str, df_all: pd.DataFrame):
    """保存累计总表（All_Full_Data + 每个产品sheet）"""
    Path(os.path.dirname(accumulated_path)).mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(accumulated_path, engine="openpyxl") as writer:
        df_all.sort_index().to_excel(writer, sheet_name="All_Full_Data", index=True)
        for epic, g in df_all.groupby("Epic"):
            sheet_name = safe_sheet_name(EPIC_TO_NAME.get(epic, epic))
            g.sort_index().to_excel(writer, sheet_name=sheet_name, index=True)


def fetch_yesterday_and_accumulate(epic_list, accumulated_excel_path: str):
    """
    ✅ 核心：只抓昨天伦敦日历日的数据，然后合并进累计总表
    返回：累计总表路径（供Step3用）
    """
    start_london, end_london, start_bj, end_bj, start_str, end_str = london_yesterday_range_to_beijing_api_strings()

    print("======================================")
    print("✅ Step 1/5: IG 增量抓取（只抓昨天 London day）并写入累计总表")
    print(f"📅 London Range : {start_london.strftime('%Y-%m-%d %H:%M:%S %Z')} -> {end_london.strftime('%Y-%m-%d %H:%M:%S %Z')}")
    print(f"📅 Beijing Range: {start_bj.strftime('%Y-%m-%d %H:%M:%S %Z')} -> {end_bj.strftime('%Y-%m-%d %H:%M:%S %Z')}")
    print(f"📤 API传入（北京时间字符串）: start={start_str}, end={end_str}")
    print(f"📄 Accumulated Excel: {accumulated_excel_path}")
    print("======================================\n")

    retryer = Retrying(wait=wait_exponential(), retry=retry_if_exception_type(ApiExceededException))
    ig_service = IGService(
        IGConfig.username,
        IGConfig.password,
        IGConfig.api_key,
        IGConfig.acc_type,
        retryer=retryer,
        use_rate_limiter=True,
    )

    df_existing = load_accumulated_excel(accumulated_excel_path)
    if df_existing is None:
        print("ℹ️ 累计总表不存在，将新建。")
    else:
        print(f"✅ 已读取累计总表：{len(df_existing)} 条记录")

    new_chunks = []

    try:
        ig_service.create_session()
        print("✅ IG Session Created Successfully")

        for i, epic in enumerate(epic_list, 1):
            product_name = EPIC_TO_NAME.get(epic, epic)
            print(f"\n--- Fetching {i}/{len(epic_list)}: {product_name} ({epic}) ---")

            df_1h = fetch_data_by_resolution(ig_service, epic, "1h", start_str, end_str)
            df_30min = fetch_data_by_resolution(ig_service, epic, "30Min", start_str, end_str)

            if not df_30min.empty:
                df_30min = df_30min[
                    (df_30min.index.hour >= HALF_HOUR_RESOLUTION_RANGE["start_hour"])
                    & (df_30min.index.hour <= HALF_HOUR_RESOLUTION_RANGE["end_hour"])
                ]

            df_list = []
            if not df_1h.empty:
                df_list.append(df_1h)
            if not df_30min.empty:
                df_list.append(df_30min)

            if not df_list:
                print("⚠️ 本产品昨天无数据，跳过")
                continue

            df_combined = pd.concat(df_list).sort_index()
            df_combined["Product Name"] = product_name
            df_combined["Epic"] = epic

            print(f"✅ 昨天新增记录数：{len(df_combined)}")
            new_chunks.append(df_combined)

        if not new_chunks:
            print("\n⚠️ 昨天所有产品都没拿到数据：累计总表不更新。")
            return accumulated_excel_path

        df_new_all = pd.concat(new_chunks).sort_index()

        # 合并到累计总表
        if df_existing is None or df_existing.empty:
            df_merged = df_new_all
        else:
            df_merged = pd.concat([df_existing, df_new_all]).sort_index()
            # 去重：同一个(时间, Epic, Resolution)出现重复时，保留最后一次
            df_merged = df_merged.reset_index()
            df_merged = df_merged.drop_duplicates(subset=["DateTime (London)", "Epic", "Resolution"], keep="last")
            df_merged = df_merged.set_index("DateTime (London)").sort_index()

        print(f"\n✅ 合并后累计总表记录数：{len(df_merged)}（新增 {len(df_new_all)}）")
        save_accumulated_excel(accumulated_excel_path, df_merged)
        print(f"💾 已更新累计总表 -> {accumulated_excel_path}\n")

        return accumulated_excel_path

    finally:
        try:
            ig_service.logout()
        except Exception:
            pass
        print("🔚 Session Closed\n")


# =============================================================================
# 3) Step 2 - 更新模板日期（8个sheet）(你的原逻辑不变)
# =============================================================================

def update_template_dates_uk(TARGET_FILE: str, OUTPUT_FILE: str):
    uk_now = datetime.now(TZ_LONDON)
    today_str = uk_now.strftime("%Y/%m/%d")
    yesterday_str = (uk_now - timedelta(days=1)).strftime("%Y/%m/%d")

    print("======================================")
    print("✅ Step 2/5: 更新模板日期（英国时间 Europe/London）")
    print(f"🧾 Read Template : {TARGET_FILE}")
    print(f"🧾 Write Updated : {OUTPUT_FILE}")
    print("📌 当前英国时间:", uk_now.strftime("%Y-%m-%d %H:%M:%S %Z"))
    print("📅 Yesterday (UK):", yesterday_str)
    print("📅 Today     (UK):", today_str)
    print("======================================\n")

    if not os.path.exists(TARGET_FILE):
        raise FileNotFoundError(f"❌ 找不到模板文件：{TARGET_FILE}")

    wb = load_workbook(TARGET_FILE)

    def update_time_sheet_cell(cell_value, new_date_str):
        if not isinstance(cell_value, str):
            return None
        if "-" not in cell_value or "时Close" not in cell_value:
            return None
        parts = cell_value.split("-", 1)
        suffix = parts[1]
        return f"{new_date_str}-{suffix}"

    def normalize_change_sheet_date_cell(cell_value):
        if cell_value is None:
            return None
        if isinstance(cell_value, datetime):
            return cell_value.strftime("%Y/%m/%d")
        if isinstance(cell_value, str):
            s = cell_value.strip()
            if s.count("/") == 2 and len(s) >= 10:
                return s[:10]
            if len(s) >= 10 and s[4] == "-" and s[7] == "-":
                return s[:10].replace("-", "/")
        return None

    print("✅ Step 2.1: 更新 time_sheets (A3-A14)")
    for sheet_name in TIME_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        for row in range(3, 9):
            cell = ws.cell(row=row, column=1)
            new = update_time_sheet_cell(cell.value, yesterday_str)
            if new is not None:
                cell.value = new

        for row in range(9, 15):
            cell = ws.cell(row=row, column=1)
            new = update_time_sheet_cell(cell.value, today_str)
            if new is not None:
                cell.value = new

    print("\n✅ Step 2.2: 更新 change_sheets (A3/A9)")
    for sheet_name in CHANGE_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        c3 = ws.cell(row=3, column=1)
        if normalize_change_sheet_date_cell(c3.value) is not None:
            c3.value = yesterday_str

        c9 = ws.cell(row=9, column=1)
        if normalize_change_sheet_date_cell(c9.value) is not None:
            c9.value = today_str

    wb.save(OUTPUT_FILE)
    print(f"\n🎉 Step 2 完成：已保存 -> {OUTPUT_FILE}\n")
    return yesterday_str, today_str


# =============================================================================
# 4) Step 3 - 对“累计总表”进行筛选（你的原逻辑不变）
# =============================================================================

def filter_historical_data_full_to_template_times(input_excel_path: str, output_excel_path: str):
    print("======================================")
    print("✅ Step 3/5: 从累计总表筛选出模板需要的时间点")
    print(f"🧾 Read  Accumulated Excel : {input_excel_path}")
    print(f"🧾 Write Filtered          : {output_excel_path}")
    print(f"🔧 Rule 1h hours            : {sorted(list(FILTER_1H_HOURS))}")
    print(f"🔧 Rule 30Min times         : {sorted(list(FILTER_30MIN_TIMES))}")
    print("======================================\n")

    df = pd.read_excel(
        input_excel_path,
        sheet_name="All_Full_Data",
        index_col="DateTime (London)",
        parse_dates=["DateTime (London)"],
    )
    print(f"✅ 成功读取累计数据：{len(df)} 条记录")

    filtered_data = []
    for product_name, group in df.groupby("Product Name"):
        group_1h = group[group["Resolution"] == "1h"].copy()
        group_30min = group[group["Resolution"] == "30Min"].copy()

        group_1h_filtered = group_1h[group_1h.index.hour.isin(FILTER_1H_HOURS)]

        group_30min["time_str"] = group_30min.index.strftime("%H:%M")
        group_30min_filtered = group_30min[group_30min["time_str"].isin(FILTER_30MIN_TIMES)]
        group_30min_filtered = group_30min_filtered.drop(columns=["time_str"], errors="ignore")

        group_filtered = pd.concat([group_1h_filtered, group_30min_filtered]).sort_index()
        filtered_data.append(group_filtered)

    if not filtered_data:
        raise RuntimeError("❌ 筛选结果为空，无法继续")

    df_filtered = pd.concat(filtered_data).sort_index()
    print(f"📊 最终筛选结果：{len(df_filtered)} 条")

    Path(os.path.dirname(output_excel_path)).mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_excel_path, engine="openpyxl") as writer:
        df_filtered.to_excel(writer, sheet_name="Filtered_Full_Data", index=True)
        for product_name, group in df_filtered.groupby("Product Name"):
            safe_name = product_name.replace("/", "_").replace("$", "USD").replace(" ", "_")[:31]
            group.to_excel(writer, sheet_name=safe_name, index=True)

    print(f"💾 Step 3 完成：筛选后数据已保存 -> {output_excel_path}\n")
    return output_excel_path


# =============================================================================
# 5) Step 4 - 写入模板（保持你原逻辑）
# =============================================================================

def parse_timestamp_label(label: str) -> datetime:
    label = str(label).strip()
    date_part, time_part = label.split("-")
    date_obj = datetime.strptime(date_part.strip(), "%Y/%m/%d")

    time_str = time_part.split("时")[0]
    if ":" in time_str:
        hour_str, minute_str = time_str.split(":")
        hour = int(hour_str)
        minute = int(minute_str)
    else:
        hour = int(time_str)
        minute = 0

    return date_obj.replace(hour=hour, minute=minute, second=0, microsecond=0)


def fill_template_with_close_data(source_file: str, template_file: str, output_file: str):
    print("======================================")
    print("✅ Step 4/5: 把筛选后的 Close 数据写入模板（05/07/15/20时）")
    print(f"🧾 Read  Filtered Data Excel : {source_file}")
    print(f"🧾 Read  Updated Template    : {template_file}")
    print(f"🧾 Write Filled Output       : {output_file}")
    print("======================================\n")

    if not os.path.exists(source_file):
        raise FileNotFoundError(f"源数据文件不存在：{source_file}")
    if not os.path.exists(template_file):
        raise FileNotFoundError(f"模板文件不存在：{template_file}")

    product_data = {}
    for header_name, sheet_name in PRODUCT_SHEET_MAP.items():
        try:
            df = pd.read_excel(source_file, sheet_name=sheet_name)
        except Exception:
            continue

        if "DateTime (London)" not in df.columns or "Close" not in df.columns:
            continue

        df["DateTime (London)"] = pd.to_datetime(df["DateTime (London)"])
        df = df.set_index("DateTime (London)")
        keep_cols = ["Close"]
        if "Resolution" in df.columns:
            keep_cols.append("Resolution")
        product_data[header_name] = df[keep_cols]

    wb = load_workbook(template_file)

    for sheet_name in TIME_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column

        col_to_product = {}
        for col in range(2, max_col + 1):
            header = ws.cell(row=1, column=col).value
            if header and header in PRODUCT_SHEET_MAP:
                col_to_product[col] = header

        for row in range(3, max_row + 1):
            label = ws.cell(row=row, column=1).value
            if not label:
                continue

            try:
                ts = parse_timestamp_label(label)
            except Exception:
                continue

            for col, product_header in col_to_product.items():
                dfp = product_data.get(product_header)
                if dfp is None:
                    continue

                if ts in dfp.index:
                    close_value = dfp.loc[ts, "Close"]
                    if hasattr(close_value, "iloc"):
                        close_value = close_value.iloc[0]
                    close_value = float(close_value)

                    # 18:30 <-> 18:00 互补
                    if ts.hour == 18 and ts.minute == 30:
                        ts_1800 = ts.replace(minute=0)
                        if ts_1800 in dfp.index:
                            close_value = float(dfp.loc[ts_1800, "Close"])
                    if ts.hour == 18 and ts.minute == 0:
                        ts_1830 = ts.replace(minute=30)
                        if ts_1830 in dfp.index:
                            close_value = float(dfp.loc[ts_1830, "Close"])

                    ws.cell(row=row, column=col).value = close_value
                else:
                    # 20:00 用 19:00
                    if ts.hour == 20 and ts.minute == 0:
                        ts_1900 = ts.replace(hour=19, minute=0)
                        if ts_1900 in dfp.index:
                            close_value = dfp.loc[ts_1900, "Close"]
                            if hasattr(close_value, "iloc"):
                                close_value = close_value.iloc[0]
                            ws.cell(row=row, column=col).value = float(close_value)

    wb.save(output_file)
    print(f"\n🎉 Step 4 完成：已保存为 -> {output_file}\n")
    return output_file


# =============================================================================
# 6) Step 5 - 发邮件（可选，保持你原逻辑）
# =============================================================================

def send_gmail_with_attachment(send_usr, send_pwd, receive_usr_list, attachment_path, email_title, content, email_server="smtp.gmail.com", email_port=587):
    print("======================================")
    print("✅ Step 5/5: 发送 Gmail 邮件（含附件）")
    print(f"📨 From: {send_usr}")
    print(f"📨 To  : {', '.join(receive_usr_list)}")
    print(f"📎 Attachment: {attachment_path}")
    print("======================================\n")


    msg = MIMEMultipart()
    msg["Subject"] = email_title
    msg["From"] = send_usr
    msg["To"] = ", ".join(receive_usr_list)
    msg.attach(MIMEText(content, "plain", "utf-8"))

    if os.path.exists(attachment_path):
        with open(attachment_path, "rb") as f:
            attachment = MIMEApplication(f.read(), _subtype="xlsx")
            attachment.add_header("Content-Disposition", "attachment", filename=os.path.basename(attachment_path))
            msg.attach(attachment)
    else:
        print(f"❌ 附件不存在：{attachment_path}")
        return

    try:
        smtp = SMTP(email_server, email_port, timeout=30)
        smtp.starttls()
        smtp.login(send_usr, send_pwd)
        smtp.sendmail(send_usr, receive_usr_list, msg.as_string())
        smtp.quit()
        print("✅ Gmail邮件（含附件）发送成功！\n")
    except Exception as e:
        print(f"❌ 发送失败：{str(e)}\n")


# =============================================================================
# 7) 主程序：一键跑完（✅ 已删除 days=2）
# =============================================================================

def main():
    print("\n" + "=" * 70)
    print("🚀 一键跑完主程序启动（增量累计版：只抓昨天）")
    print("=" * 70)

    epic_list = list(EPIC_TO_NAME.keys())

    # Step 1: 只抓昨天，并累计到一个总表
    accumulated_excel_path = fetch_yesterday_and_accumulate(
        epic_list=epic_list,
        accumulated_excel_path=ACCUMULATED_EXCEL_FILE
    )

    # Step 2: 更新模板日期
    update_template_dates_uk(TEMPLATE_FILE, UPDATED_TEMPLATE_FILE)

    # Step 3: 从累计总表筛选
    output_dir_abs = os.path.dirname(accumulated_excel_path)
    filtered_excel_abs = os.path.join(output_dir_abs, FILTERED_DATA_EXCEL_NAME)
    filter_historical_data_full_to_template_times(accumulated_excel_path, filtered_excel_abs)

    # Step 4: 填充模板
    fill_template_with_close_data(
        source_file=filtered_excel_abs,
        template_file=UPDATED_TEMPLATE_FILE,
        output_file=FILLED_OUTPUT_FILE,
    )

    # Step 5: 发邮件（可选）
    if SEND_EMAIL:
        gmail_config = get_gmail_config()
        email_title = f"Excel数据附件 - {datetime.now().strftime('%Y%m%d')} - 变化率表格"
        content = "这是用Python脚本发送的邮件(变化率表格)，附带Excel数据附件，请查收！"

        send_gmail_with_attachment(
            send_usr=gmail_config.send_usr,
            send_pwd=gmail_config.send_pwd,
            receive_usr_list=gmail_config.receive_usr_list,
            attachment_path=FILLED_OUTPUT_FILE,
            email_title=email_title,
            content=content,
            email_server=gmail_config.email_server,
            email_port=gmail_config.email_port,
        )
    else:        print("ℹ️ Step5 已跳过（SEND_EMAIL = False）\n")

    print("=" * 70)
    print("🎉 全流程完成！")
    print(f"📄 累计总表（核心） : {ACCUMULATED_EXCEL_FILE}")
    print(f"📄 筛选后数据Excel  : {filtered_excel_abs}")
    print(f"📄 更新后模板       : {UPDATED_TEMPLATE_FILE}")
    print(f"📄 最终填好数据表   : {FILLED_OUTPUT_FILE}")
    print("=" * 70 + "\n")


if __name__ == "__main__":
    main()
