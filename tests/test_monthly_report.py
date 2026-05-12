# -*- coding: utf-8 -*-
import tempfile
import unittest
from datetime import date
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook


TIME_SHEETS = ["05时", "07时", "15时", "20时"]
CHANGE_SHEETS = ["05变化率", "07变化率", "15变化率", "20变化率"]
TIME_ROWS = {
    "05时": ["05", "18", "18:30", "19", "19", "20"],
    "07时": ["07", "18", "18:30", "19", "19", "20"],
    "15时": ["15", "18", "18:30", "19", "19", "20"],
    "20时": ["18", "18", "18:30", "19", "19", "20"],
}
CHANGE_ROWS = {
    "05变化率": ["05'时-05时", "18'时-05'时", "18:30'-05'时", "19'时-05'时", "19'时-05'时", "20'时-05'时"],
    "07变化率": ["07'时-07时", "18'时-07'时", "18:30'-07'时", "19'时-07'时", "19'时-07'时", "20'时-07'时"],
    "15变化率": ["15'时-15时", "18'时-15'时", "18:30'-15'时", "19'时-15'时", "19'时-15'时", "20'时-15'时"],
    "20变化率": ["18'时-19时", "18'时-20时", "18:30'时-20时", "19'时-19时", "19'时-20时", "20'时-20时"],
}


def _date_text(value):
    return value.strftime("%Y/%m/%d")


def _create_daily_workbook(path, dates):
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name in TIME_SHEETS:
        ws = wb.create_sheet(sheet_name)
        ws["A2"] = "Timestamp"
        ws["B1"] = "US500"
        ws["B2"] = "Close"
        ws["C2"] = "Change"
        ws["D1"] = "HK50"
        ws["D2"] = "Close"
        ws["E2"] = "Change"
        ws["F1"] = "HK50-US500"
        ws["F2"] = "Change"

        row = 3
        for day in dates:
            for offset, time_label in enumerate(TIME_ROWS[sheet_name]):
                ws.cell(row, 1).value = f"{_date_text(day)}-{time_label}时Close"
                ws.cell(row, 2).value = 7000 + day.day * 10 + offset
                ws.cell(row, 4).value = 26000 + day.day * 10 + offset
                row += 1

        # Rows 9-14 in each daily workbook hold the formula pattern copied by the report builder.
        if sheet_name == "20时":
            row_formulas = [
                ("=B9/B7-1", "=D9/D7-1"),
                ("=B10/B8-1", "=D10/D8-1"),
                ("=B11/B8-1", "=D11/D8-1"),
                ("=B12/B7-1", "=D12/D7-1"),
                ("=B13/B8-1", "=D13/D8-1"),
                ("=B14/B8-1", "=D14/D8-1"),
            ]
        else:
            row_formulas = [
                ("=B9/B3-1", "=D9/D3-1"),
                ("=B10/$B$9-1", "=D10/$D$9-1"),
                ("=B11/$B$9-1", "=D11/$D$9-1"),
                ("=B12/$B$9-1", "=D12/$D$9-1"),
                ("=B13/$B$9-1", "=D13/$D$9-1"),
                ("=B14/$B$9-1", "=D14/$D$9-1"),
            ]
        for idx, (us_formula, hk_formula) in enumerate(row_formulas, start=9):
            ws.cell(idx, 3).value = us_formula
            ws.cell(idx, 5).value = hk_formula
            ws.cell(idx, 6).value = f"=E{idx}-C{idx}"

    for sheet_name in CHANGE_SHEETS:
        ws = wb.create_sheet(sheet_name)
        ws["A2"] = "Date"
        ws["B2"] = "Time"
        ws["C1"] = "US500"
        ws["C2"] = "Change"
        ws["D1"] = "HK50"
        ws["D2"] = "Change"
        ws["E1"] = "HK50-US500"
        ws["E2"] = "Change"
        time_sheet = sheet_name.replace("变化率", "时")
        row = 3
        for day in dates:
            for offset, label in enumerate(CHANGE_ROWS[sheet_name]):
                ws.cell(row, 1).value = _date_text(day) if offset == 0 else None
                ws.cell(row, 2).value = label
                ws.cell(row, 3).value = f"='{time_sheet}'!C{row}"
                ws.cell(row, 4).value = f"='{time_sheet}'!E{row}"
                ws.cell(row, 5).value = f"='{time_sheet}'!F{row}"
                row += 1

    wb.save(path)


class MonthlyReportTests(unittest.TestCase):
    def test_collects_all_py_final_workbook_names(self):
        from monthly_report import collect_daily_workbooks

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            historical = root / "historical_data_20260509_040500"
            historical.mkdir()
            expected = historical / "IG变化率_20260509.xlsx"
            ignored = historical / "IG变化率_202605_公式版.xlsx"
            expected.write_bytes(b"daily")
            ignored.write_bytes(b"monthly")

            self.assertEqual([expected], collect_daily_workbooks(str(root)))

    def test_builds_formula_report_with_previous_month_reference(self):
        from monthly_report import build_monthly_formula_report

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            first = root / "IG变化率_20260501.xlsx"
            second = root / "IG变化率_20260502.xlsx"
            latest = root / "IG变化率_20260509.xlsx"
            _create_daily_workbook(first, [date(2026, 4, 29), date(2026, 4, 30)])
            _create_daily_workbook(second, [date(2026, 4, 30), date(2026, 5, 1)])
            _create_daily_workbook(latest, [date(2026, 5, 7), date(2026, 5, 8)])

            output = root / "IG变化率_202605_公式版.xlsx"
            result = build_monthly_formula_report(
                output_root_dir=str(root),
                output_file=str(output),
                report_month="202605",
                input_files=[str(first), str(second), str(latest)],
            )

            self.assertEqual(str(output), result)
            wb = load_workbook(output, data_only=False)
            self.assertEqual(TIME_SHEETS + CHANGE_SHEETS, wb.sheetnames[:8])
            self.assertEqual("hidden", wb["FormulaRefs"].sheet_state)
            self.assertEqual("2026/04/30-05时Close", wb["05时"]["A3"].value)
            self.assertEqual("2026/05/01-05时Close", wb["05时"]["A9"].value)
            self.assertEqual("=B3/FormulaRefs!B2-1", wb["05时"]["C3"].value)
            self.assertEqual("=B9/B3-1", wb["05时"]["C9"].value)
            self.assertEqual("=B10/$B$9-1", wb["05时"]["C10"].value)
            self.assertEqual("=B9/B7-1", wb["20时"]["C9"].value)
            self.assertEqual("='20时'!C9", wb["20变化率"]["C9"].value)

    def test_send_gmail_with_attachments_sends_all_files(self):
        from monthly_report import send_gmail_with_attachments

        class FakeSMTP:
            instances = []

            def __init__(self, server, port, timeout=None):
                self.server = server
                self.port = port
                self.timeout = timeout
                self.logged_in = None
                self.sent = None
                self.quit_called = False
                FakeSMTP.instances.append(self)

            def starttls(self):
                self.tls_started = True

            def login(self, username, password):
                self.logged_in = (username, password)

            def sendmail(self, sender, recipients, message):
                self.sent = (sender, recipients, message)

            def quit(self):
                self.quit_called = True

        FakeSMTP.instances = []
        with tempfile.TemporaryDirectory() as tmp:
            first = Path(tmp) / "first.xlsx"
            second = Path(tmp) / "second.xlsx"
            first.write_bytes(b"one")
            second.write_bytes(b"two")

            with patch("monthly_report.SMTP", FakeSMTP):
                sent = send_gmail_with_attachments(
                    send_usr="sender@example.com",
                    send_pwd="password",
                    receive_usr_list=["to@example.com"],
                    attachment_paths=[str(first), str(second)],
                    email_title="subject",
                    content="body",
                )

        self.assertTrue(sent)
        smtp = FakeSMTP.instances[0]
        self.assertEqual(("sender@example.com", "password"), smtp.logged_in)
        self.assertEqual("sender@example.com", smtp.sent[0])
        self.assertEqual(["to@example.com"], smtp.sent[1])
        self.assertIn("first.xlsx", smtp.sent[2])
        self.assertIn("second.xlsx", smtp.sent[2])
        self.assertTrue(smtp.quit_called)


if __name__ == "__main__":
    unittest.main()
